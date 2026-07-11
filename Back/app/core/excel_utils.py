from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO

def setup_excel_headers(ws, headers: list[str]):
    """Applies styles to headers and adjusts basic widths."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
        # Adjust column width
        ws.column_dimensions[get_column_letter(col_num)].width = max(len(header) + 5, 15)

def export_generic_excel(title: str, headers: list[str], data: list[list]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    
    setup_excel_headers(ws, headers)
    
    for row in data:
        ws.append(row)
        
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def export_plantilla_excel(title: str, headers: list[str], example_rows: list[list]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    
    setup_excel_headers(ws, headers)
    
    for row in example_rows:
        ws.append(row)
        
    # Agrega un mensaje explicativo en la última columna si es necesario
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

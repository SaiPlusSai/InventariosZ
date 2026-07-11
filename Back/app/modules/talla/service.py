from io import BytesIO
from openpyxl import Workbook, load_workbook
from fastapi import UploadFile
from app.modules.talla.schemas import PreviaImportacionResponse, FilaPrevia, TallaCreate
from app.core.exceptions import RegistroActivoNoPuedeEliminarseException
from sqlalchemy.orm import Session

from app.modules.talla.constants import (
    TALLA_NO_EXISTE,
    TALLA_YA_EXISTE,
)
from app.modules.talla.exceptions import (
    TallaNoEncontradaException,
)
from app.modules.talla.models import Talla
from app.modules.talla.repository import TallaRepository
from app.modules.talla.schemas import (
    TallaCreate,
    TallaUpdate,
)


class TallaService:

    def __init__(self):
        self.repository = TallaRepository()

    
    def get_papelera(self, db: Session) -> list[Talla]:
        return self.repository.get_papelera(db)

    def get_dependencias(self, db: Session, id: int) -> dict:
        return self.repository.get_dependencias(db, id)

    def desactivar(self, db: Session, id: int):
        item = self.repository.get_by_id(db, id)
        if item:
            item.estado = False
            from datetime import datetime
            item.deleted_at = datetime.now()
            db.commit()
            db.refresh(item)
        return item

    def recuperar(self, db: Session, id: int):
        item = self.repository.get_by_id_papelera(db, id)
        if item:
            conflicto = self.repository.get_by_nombre(db, item.nombre)
            if conflicto and conflicto.id != item.id:
                from app.core.exceptions import RecuperacionConflictivaException
                raise RecuperacionConflictivaException(
                    f"No se puede recuperar. Ya existe una talla activa con el nombre '{item.nombre}'."
                )

            item.estado = True
            item.deleted_at = None
            db.commit()
            db.refresh(item)
        return item

    
    def exportar_excel(self, db: Session) -> BytesIO:
        items = self.repository.get_all(db)
        wb = Workbook()
        ws = wb.active
        ws.title = "Talla"
        headers = ["ID", "Nombre", "Descripción", "Estado"]
        ws.append(headers)
        for item in items:
            ws.append([
                item.id,
                item.nombre,
                getattr(item, 'descripcion', ''),
                "Activo" if item.estado else "Inactivo"
                
            ])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def generar_plantilla_importacion(self) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla Talla"
        headers = ["Nombre", "Descripción"]
        ws.append(headers)
        ws.append(["Ejemplo Talla", "Descripción de ejemplo"])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    async def previa_importacion(self, db: Session, file: UploadFile) -> PreviaImportacionResponse:
        content = await file.read()
        wb = load_workbook(filename=BytesIO(content), data_only=True)
        ws = wb.active

        filas = []
        validos = 0
        errores = 0

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue
            
            nombre, descripcion = (list(row)[:2] + [None]*2)[:2]
            
            fila_errores = []
            nombre_str = str(nombre).strip() if nombre is not None else None
            desc_str = str(descripcion).strip() if descripcion is not None else None
            

            if not nombre_str: 
                fila_errores.append("Nombre es obligatorio.")
            else:
                existente = self.repository.get_by_nombre(db, nombre_str)
                if existente:
                    fila_errores.append(f"Talla '{nombre_str}' ya existe.")
            
            

            es_valido = len(fila_errores) == 0
            if es_valido:
                validos += 1
            else:
                errores += 1

            filas.append(FilaPrevia(
                fila=idx,
                nombre=nombre_str,
                descripcion=desc_str,
                
                valido=es_valido,
                errores=fila_errores
            ))

        return PreviaImportacionResponse(total=len(filas), validos=validos, errores=errores, filas=filas)

    def confirmar_importacion(self, db: Session, filas: list[FilaPrevia]):
        try:
            creados = 0
            for fila in filas:
                if not fila.valido: continue
                payload = TallaCreate(nombre=fila.nombre, descripcion=fila.descripcion or '')
                self.create(db, payload)
                creados += 1
            db.commit()
            return {"success": True, "creados": creados}
        except Exception as e:
            db.rollback()
            raise e

    def get_all(
        self,
        db: Session,
    ) -> list[Talla]:

        return self.repository.get_all(db)

    def get_by_id(
        self,
        db: Session,
        talla_id: int,
    ) -> Talla | None:

        return self.repository.get_by_id(db, talla_id)

    def create(
        self,
        db: Session,
        data: TallaCreate,
    ) -> Talla:

        talla_existente = self.repository.get_by_nombre_any_state(
            db,
            data.nombre,
        )

        if talla_existente:
            if talla_existente.estado:
                from app.core.exceptions import RegistroYaExisteException
                raise RegistroYaExisteException("La talla ya existe.")
            else:
                from app.core.exceptions import RegistroEnPapeleraException
                raise RegistroEnPapeleraException(
                    message=f"La talla '{data.nombre}' existe en la Papelera.",
                    id_registro=talla_existente.id,
                    tipo_registro="talla"
                )

        nueva_talla = Talla(
            nombre=data.nombre,
            orden=data.orden,
        )

        return self.repository.create(
            db,
            nueva_talla,
        )

    def update(
        self,
        db: Session,
        talla_id: int,
        data: TallaUpdate,
    ) -> Talla:

        talla = self.repository.get_by_id(
            db,
            talla_id,
        )

        if not talla:
            raise TallaNoEncontradaException(TALLA_NO_EXISTE)

        if data.nombre is not None:
            talla.nombre = data.nombre

        if data.orden is not None:
            talla.orden = data.orden

        if data.estado is not None:
            talla.estado = data.estado

        return self.repository.update(
            db,
            talla,
        )

    def delete(
        self,
        db: Session,
        talla_id: int,
    ) -> None:

        talla = self.repository.get_by_id(
            db,
            talla_id,
        )
        if not talla:
            talla = self.repository.get_by_id_papelera(db, talla_id if 'talla_id' in locals() else id)

        if not talla:
            raise TallaNoEncontradaException(
                TALLA_NO_EXISTE
            )

        if locals().get('item') and getattr(locals()['item'], 'estado', False) or (locals().get('talla') and getattr(locals().get('talla'), 'estado', False)):
            raise RegistroActivoNoPuedeEliminarseException('No se puede eliminar físicamente un registro activo. Envíelo a la papelera primero.')
        if talla.estado == True:
            raise RegistroActivoNoPuedeEliminarseException('No se puede eliminar un registro activo.')

        self.repository.delete(
            db,
            talla,
        )

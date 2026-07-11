from io import BytesIO
from openpyxl import Workbook, load_workbook
from fastapi import UploadFile
from app.modules.material.schemas import PreviaImportacionResponse, FilaPrevia, MaterialCreate
from app.core.exceptions import RegistroActivoNoPuedeEliminarseException
from sqlalchemy.orm import Session

from app.modules.material.constants import (
    MATERIAL_NO_EXISTE,
    MATERIAL_YA_EXISTE,
)
from app.modules.material.exceptions import (
    MaterialNoEncontradoException,
)
from app.modules.material.models import Material
from app.modules.material.repository import MaterialRepository
from app.modules.material.schemas import (
    MaterialCreate,
    MaterialUpdate,
)


class MaterialService:

    def __init__(self):
        self.repository = MaterialRepository()

    
    def get_papelera(self, db: Session) -> list[Material]:
        return self.repository.get_papelera(db)

    def get_dependencias(self, db: Session, id: int) -> dict:
        return self.repository.get_dependencias(db, id)

    def desactivar(self, db: Session, id: int):
        item = self.repository.get_by_id(db, id)
        if item:
            item.estado = False
            from datetime import datetime
            item.deleted_at = datetime.now()
            db.commit()
            db.refresh(item)
        return item

    def recuperar(self, db: Session, id: int):
        item = self.repository.get_by_id_papelera(db, id)
        if item:
            conflicto = self.repository.get_by_nombre(db, item.nombre)
            if conflicto and conflicto.id != item.id:
                from app.core.exceptions import RecuperacionConflictivaException
                raise RecuperacionConflictivaException(
                    f"No se puede recuperar. Ya existe un material activo con el nombre '{item.nombre}'."
                )

            item.estado = True
            item.deleted_at = None
            db.commit()
            db.refresh(item)
        return item

    
    def exportar_excel(self, db: Session) -> BytesIO:
        items = self.repository.get_all(db)
        wb = Workbook()
        ws = wb.active
        ws.title = "Material"
        headers = ["ID", "Nombre", "Descripción", "Estado"]
        ws.append(headers)
        for item in items:
            ws.append([
                item.id,
                item.nombre,
                getattr(item, 'descripcion', ''),
                "Activo" if item.estado else "Inactivo"
                
            ])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def generar_plantilla_importacion(self) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla Material"
        headers = ["Nombre", "Descripción"]
        ws.append(headers)
        ws.append(["Ejemplo Material", "Descripción de ejemplo"])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    async def previa_importacion(self, db: Session, file: UploadFile) -> PreviaImportacionResponse:
        content = await file.read()
        wb = load_workbook(filename=BytesIO(content), data_only=True)
        ws = wb.active

        filas = []
        validos = 0
        errores = 0

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue
            
            nombre, descripcion = (list(row)[:2] + [None]*2)[:2]
            
            fila_errores = []
            nombre_str = str(nombre).strip() if nombre is not None else None
            desc_str = str(descripcion).strip() if descripcion is not None else None
            

            if not nombre_str: 
                fila_errores.append("Nombre es obligatorio.")
            else:
                existente = self.repository.get_by_nombre(db, nombre_str)
                if existente:
                    fila_errores.append(f"Material '{nombre_str}' ya existe.")
            
            

            es_valido = len(fila_errores) == 0
            if es_valido:
                validos += 1
            else:
                errores += 1

            filas.append(FilaPrevia(
                fila=idx,
                nombre=nombre_str,
                descripcion=desc_str,
                
                valido=es_valido,
                errores=fila_errores
            ))

        return PreviaImportacionResponse(total=len(filas), validos=validos, errores=errores, filas=filas)

    def confirmar_importacion(self, db: Session, filas: list[FilaPrevia]):
        try:
            creados = 0
            for fila in filas:
                if not fila.valido: continue
                payload = MaterialCreate(nombre=fila.nombre, descripcion=fila.descripcion or '')
                self.create(db, payload)
                creados += 1
            db.commit()
            return {"success": True, "creados": creados}
        except Exception as e:
            db.rollback()
            raise e

    def get_all(
        self,
        db: Session,
    ) -> list[Material]:

        return self.repository.get_all(db)

    def get_by_id(
        self,
        db: Session,
        material_id: int,
    ) -> Material | None:

        return self.repository.get_by_id(db, material_id)

    def create(
        self,
        db: Session,
        data: MaterialCreate,
    ) -> Material:

        material_existente = self.repository.get_by_nombre_any_state(
            db,
            data.nombre,
        )

        if material_existente:
            if material_existente.estado:
                from app.core.exceptions import RegistroYaExisteException
                raise RegistroYaExisteException("El material ya existe.")
            else:
                from app.core.exceptions import RegistroEnPapeleraException
                raise RegistroEnPapeleraException(
                    message=f"El material '{data.nombre}' existe en la Papelera.",
                    id_registro=material_existente.id,
                    tipo_registro="material"
                )

        nuevo_material = Material(
            nombre=data.nombre,
            descripcion=data.descripcion,
        )

        return self.repository.create(
            db,
            nuevo_material,
        )

    def update(
        self,
        db: Session,
        material_id: int,
        data: MaterialUpdate,
    ) -> Material:

        material = self.repository.get_by_id(
            db,
            material_id,
        )

        if not material:
            raise MaterialNoEncontradoException(MATERIAL_NO_EXISTE)

        if data.nombre is not None:
            material.nombre = data.nombre

        if data.descripcion is not None:
            material.descripcion = data.descripcion

        if data.estado is not None:
            material.estado = data.estado

        return self.repository.update(
            db,
            material,
        )

    def delete(
        self,
        db: Session,
        material_id: int,
    ) -> None:

        material = self.repository.get_by_id(
            db,
            material_id,
        )
        if not material:
            material = self.repository.get_by_id_papelera(db, material_id if 'material_id' in locals() else id)

        if not material:
            raise MaterialNoEncontradoException(
                MATERIAL_NO_EXISTE
            )

        if locals().get('item') and getattr(locals()['item'], 'estado', False) or (locals().get('material') and getattr(locals().get('material'), 'estado', False)):
            raise RegistroActivoNoPuedeEliminarseException('No se puede eliminar físicamente un registro activo. Envíelo a la papelera primero.')
        if material.estado == True:
            raise RegistroActivoNoPuedeEliminarseException('No se puede eliminar un registro activo.')

        self.repository.delete(
            db,
            material,
        )

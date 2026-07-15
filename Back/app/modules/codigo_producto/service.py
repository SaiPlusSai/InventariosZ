from app.core.exceptions import RegistroActivoNoPuedeEliminarseException
from sqlalchemy.orm import Session
from io import BytesIO
from openpyxl import Workbook, load_workbook
from fastapi import UploadFile
from app.core.excel_utils import export_generic_excel, export_plantilla_excel
from app.modules.codigo_producto.schemas import FilaPrevia, PreviaImportacionResponse

from app.modules.codigo_producto.constants import (
    CODIGO_PRODUCTO_NO_EXISTE,
    CODIGO_PRODUCTO_YA_EXISTE,
)

from app.modules.codigo_producto.exceptions import (
    CodigoProductoNoEncontradoException,
    CodigoProductoYaExisteException,
)

from app.modules.codigo_producto.models import CodigoProducto
from app.modules.codigo_producto.repository import CodigoProductoRepository
from app.modules.codigo_producto.schemas import (
    CodigoProductoCreate,
    CodigoProductoUpdate,
)

from app.modules.marca.constants import MARCA_NO_EXISTE
from app.modules.marca.exceptions import MarcaNoEncontradaException
from app.modules.marca.repository import MarcaRepository


class CodigoProductoService:

    def __init__(self):
        self.repository = CodigoProductoRepository()
        self.marca_repository = MarcaRepository()

    
    def get_papelera(self, db: Session) -> list[CodigoProducto]:
        return self.repository.get_papelera(db)

    def get_dependencias(self, db: Session, id: int) -> dict:
        return self.repository.get_dependencias(db, id)

    def desactivar(self, db: Session, id: int):
        item = self.repository.get_by_id(db, id)
        if item:
            item.estado = False
            from datetime import datetime
            item.deleted_at = datetime.now()
            db.commit()
            db.refresh(item)
        return item

    def recuperar(self, db: Session, id: int):
        item = self.repository.get_by_id_papelera(db, id)
        if item:
            item.estado = True
            item.deleted_at = None
            db.commit()
            db.refresh(item)
        return item
        
    def exportar_excel(self, db: Session) -> BytesIO:
        items = self.repository.get_all(db)
        data = [
            [
                item.id,
                item.marca.nombre if item.marca else '',
                item.codigo,
                "Activo" if item.estado else "Inactivo"
            ]
            for item in items
        ]
        return export_generic_excel("Codigos_Producto", ["ID", "Marca", "Código", "Estado"], data)

    def generar_plantilla_importacion(self) -> BytesIO:
        return export_plantilla_excel(
            "Plantilla CodigoProducto", 
            ["Marca", "Código"], 
            [["Nike", "NK-001"]]
        )

    async def previa_importacion(self, db: Session, file: UploadFile) -> PreviaImportacionResponse:
        content = await file.read()
        wb = load_workbook(filename=BytesIO(content), data_only=True)
        ws = wb.active

        filas = []
        validos = 0
        errores = 0

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue
            
            marca_val, codigo_val = (list(row)[:2] + [None]*2)[:2]
            
            fila_errores = []
            marca_str = str(marca_val).strip() if marca_val is not None else None
            codigo_str = str(codigo_val).strip() if codigo_val is not None else None

            if not marca_str: 
                fila_errores.append("Marca es obligatoria.")
            if not codigo_str:
                fila_errores.append("Código es obligatorio.")
                
            marca_obj = None
            if marca_str:
                marca_obj = self.marca_repository.get_by_nombre(db, marca_str)
                if not marca_obj:
                    fila_errores.append(f"La marca '{marca_str}' no existe.")
            
            es_valido = len(fila_errores) == 0
            if es_valido:
                validos += 1
            else:
                errores += 1

            filas.append(FilaPrevia(
                fila=idx,
                marca=marca_str,
                codigo=codigo_str,
                valido=es_valido,
                errores=fila_errores
            ))

        return PreviaImportacionResponse(total=len(filas), validos=validos, errores=errores, filas=filas)

    def confirmar_importacion(self, db: Session, filas: list[FilaPrevia]):
        try:
            creados = 0
            for fila in filas:
                if not fila.valido: continue
                marca_obj = self.marca_repository.get_by_nombre(db, fila.marca)
                payload = CodigoProductoCreate(marca_id=marca_obj.id, codigo=fila.codigo)
                self.create(db, payload)
                creados += 1
            db.commit()
            return {"success": True, "creados": creados}
        except Exception as e:
            db.rollback()
            raise e

    def get_all(
        self,
        db: Session,
    ) -> list[CodigoProducto]:

        return self.repository.get_all(db)

    def get_by_id(
        self,
        db: Session,
        codigo_producto_id: int,
    ) -> CodigoProducto | None:

        return self.repository.get_by_id(
            db,
            codigo_producto_id,
        )

    def create(
        self,
        db: Session,
        data: CodigoProductoCreate,
    ) -> CodigoProducto:

        marca = self.marca_repository.get_by_id(
            db,
            data.marca_id,
        )

        if not marca:
            raise MarcaNoEncontradaException(
                MARCA_NO_EXISTE
            )

        otros = self.repository.get_all_by_codigo(db, data.codigo)
        if otros and not getattr(data, 'force', False):
            marca_conflicto_obj = self.marca_repository.get_by_id(db, otros[0].marca_id)
            marca_conflicto_nombre = marca_conflicto_obj.nombre if marca_conflicto_obj else "Otra marca"
            from app.core.exceptions import CodigoProductoOtraMarcaWarning
            raise CodigoProductoOtraMarcaWarning(
                f"El código '{data.codigo}' ya se encuentra registrado para la marca '{marca_conflicto_nombre}'.",
                codigo=data.codigo,
                marca_conflicto=marca_conflicto_nombre,
                marca_destino=marca.nombre
            )

        nuevo_codigo = CodigoProducto(
            marca_id=data.marca_id,
            codigo=data.codigo,
        )

        return self.repository.create(
            db,
            nuevo_codigo,
        )

    def update(
        self,
        db: Session,
        codigo_producto_id: int,
        data: CodigoProductoUpdate,
    ) -> CodigoProducto:

        codigo_producto = self.repository.get_by_id(
            db,
            codigo_producto_id,
        )

        if not codigo_producto:
            raise CodigoProductoNoEncontradoException(
                CODIGO_PRODUCTO_NO_EXISTE
            )

        if data.marca_id is not None:

            marca = self.marca_repository.get_by_id(
                db,
                data.marca_id,
            )

            if not marca:
                raise MarcaNoEncontradaException(
                    MARCA_NO_EXISTE
                )

            codigo_producto.marca_id = data.marca_id

        if data.codigo is not None or data.marca_id is not None:
            new_codigo = data.codigo if data.codigo is not None else codigo_producto.codigo
            new_marca_id = data.marca_id if data.marca_id is not None else codigo_producto.marca_id

            if new_codigo != codigo_producto.codigo:
                otros = self.repository.get_all_by_codigo(db, new_codigo)
                otros_filtrados = [o for o in otros if o.id != codigo_producto.id]
                if otros_filtrados and not getattr(data, 'force', False):
                    marca_conflicto_obj = self.marca_repository.get_by_id(db, otros_filtrados[0].marca_id)
                    marca_conflicto_nombre = marca_conflicto_obj.nombre if marca_conflicto_obj else "Otra marca"
                    marca_destino_obj = self.marca_repository.get_by_id(db, new_marca_id)
                    marca_destino_nombre = marca_destino_obj.nombre if marca_destino_obj else "la marca especificada"
                    from app.core.exceptions import CodigoProductoOtraMarcaWarning
                    raise CodigoProductoOtraMarcaWarning(
                        f"El código '{new_codigo}' ya se encuentra registrado para la marca '{marca_conflicto_nombre}'.",
                        codigo=new_codigo,
                        marca_conflicto=marca_conflicto_nombre,
                        marca_destino=marca_destino_nombre
                    )

            if data.codigo is not None:
                codigo_producto.codigo = data.codigo

        if data.estado is not None:
            codigo_producto.estado = data.estado

        return self.repository.update(
            db,
            codigo_producto,
        )

    def delete(
        self,
        db: Session,
        codigo_producto_id: int,
    ) -> None:

        codigo_producto = self.repository.get_by_id(
            db,
            codigo_producto_id,
        )
        if not codigo_producto:
            codigo_producto = self.repository.get_by_id_papelera(db, codigo_producto_id if 'codigo_producto_id' in locals() else id)

        if not codigo_producto:
            raise CodigoProductoNoEncontradoException(
                CODIGO_PRODUCTO_NO_EXISTE
            )

        if locals().get('item') and getattr(locals()['item'], 'estado', False) or (locals().get('codigo_producto') and getattr(locals().get('codigo_producto'), 'estado', False)):
            raise RegistroActivoNoPuedeEliminarseException('No se puede eliminar físicamente un registro activo. Envíelo a la papelera primero.')
        if codigo_producto.estado == True:
            raise RegistroActivoNoPuedeEliminarseException('No se puede eliminar un registro activo.')

        self.repository.delete(
            db,
            codigo_producto,
        )
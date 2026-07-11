from io import BytesIO
from openpyxl import Workbook, load_workbook
from fastapi import UploadFile
from app.modules.marca.schemas import PreviaImportacionResponse, FilaPrevia, MarcaCreate
from app.core.exceptions import RegistroActivoNoPuedeEliminarseException
from sqlalchemy.orm import Session
from app.modules.marca.constants import (
    MARCA_NO_EXISTE,
    MARCA_YA_EXISTE,
)

from app.modules.marca.exceptions import (
    MarcaNoEncontradaException,
)
from app.modules.marca.models import Marca
from app.modules.marca.repository import MarcaRepository
from app.modules.marca.schemas import (
    MarcaCreate,
    MarcaUpdate,
)


class MarcaService:

    def __init__(self):
        self.repository = MarcaRepository()

    
    def get_papelera(self, db: Session) -> list[Marca]:
        return self.repository.get_papelera(db)

    def get_dependencias(self, db: Session, id: int) -> dict:
        return self.repository.get_dependencias(db, id)

    def desactivar(self, db: Session, id: int):
        item = self.repository.get_by_id(db, id)
        if item:
            item.estado = False
            from datetime import datetime
            item.deleted_at = datetime.now()
            db.commit()
            db.refresh(item)
        return item

    def recuperar(self, db: Session, id: int):
        item = self.repository.get_by_id_papelera(db, id)
        if item:
            conflicto = self.repository.get_by_nombre(db, item.nombre)
            if conflicto and conflicto.id != item.id:
                from app.core.exceptions import RecuperacionConflictivaException
                raise RecuperacionConflictivaException(
                    f"No se puede recuperar. Ya existe una marca activa con el nombre '{item.nombre}'."
                )

            item.estado = True
            item.deleted_at = None
            db.commit()
            db.refresh(item)
        return item

    
    def exportar_excel(self, db: Session) -> BytesIO:
        items = self.repository.get_all(db)
        wb = Workbook()
        ws = wb.active
        ws.title = "Marca"
        headers = ["ID", "Nombre", "Descripción", "Estado"]
        ws.append(headers)
        for item in items:
            ws.append([
                item.id,
                item.nombre,
                getattr(item, 'descripcion', ''),
                "Activo" if item.estado else "Inactivo"
                
            ])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def generar_plantilla_importacion(self) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla Marca"
        headers = ["Nombre", "Descripción"]
        ws.append(headers)
        ws.append(["Ejemplo Marca", "Descripción de ejemplo"])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    async def previa_importacion(self, db: Session, file: UploadFile) -> PreviaImportacionResponse:
        content = await file.read()
        wb = load_workbook(filename=BytesIO(content), data_only=True)
        ws = wb.active

        filas = []
        validos = 0
        errores = 0

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue
            
            nombre, descripcion = (list(row)[:2] + [None]*2)[:2]
            
            fila_errores = []
            nombre_str = str(nombre).strip() if nombre is not None else None
            desc_str = str(descripcion).strip() if descripcion is not None else None
            

            if not nombre_str: 
                fila_errores.append("Nombre es obligatorio.")
            else:
                existente = self.repository.get_by_nombre(db, nombre_str)
                if existente:
                    fila_errores.append(f"Marca '{nombre_str}' ya existe.")
            
            

            es_valido = len(fila_errores) == 0
            if es_valido:
                validos += 1
            else:
                errores += 1

            filas.append(FilaPrevia(
                fila=idx,
                nombre=nombre_str,
                descripcion=desc_str,
                
                valido=es_valido,
                errores=fila_errores
            ))

        return PreviaImportacionResponse(total=len(filas), validos=validos, errores=errores, filas=filas)

    def confirmar_importacion(self, db: Session, filas: list[FilaPrevia]):
        try:
            creados = 0
            for fila in filas:
                if not fila.valido: continue
                payload = MarcaCreate(nombre=fila.nombre, descripcion=fila.descripcion or '')
                self.create(db, payload)
                creados += 1
            db.commit()
            return {"success": True, "creados": creados}
        except Exception as e:
            db.rollback()
            raise e

    def get_all(
        self,
        db: Session,
    ) -> list[Marca]:

        return self.repository.get_all(db)

    def get_by_id(
        self,
        db: Session,
        marca_id: int,
    ) -> Marca | None:

        return self.repository.get_by_id(db, marca_id)

    def create(
        self,
        db: Session,
        data: MarcaCreate,
    ) -> Marca:

        marca_existente = self.repository.get_by_nombre_any_state(
            db,
            data.nombre,
        )

        if marca_existente:
            if marca_existente.estado:
                from app.core.exceptions import RegistroYaExisteException
                raise RegistroYaExisteException("La marca ya existe.")
            else:
                from app.core.exceptions import RegistroEnPapeleraException
                raise RegistroEnPapeleraException(
                    message=f"La marca '{data.nombre}' existe en la Papelera.",
                    id_registro=marca_existente.id,
                    tipo_registro="marca"
                )

        nueva_marca = Marca(
            nombre=data.nombre,
            descripcion=data.descripcion,
        )

        return self.repository.create(
            db,
            nueva_marca,
        )

    def update(
        self,
        db: Session,
        marca_id: int,
        data: MarcaUpdate,
    ) -> Marca:

        marca = self.repository.get_by_id(
            db,
            marca_id,
        )

        if not marca:
            raise MarcaNoEncontradaException(MARCA_NO_EXISTE)

        if data.nombre is not None:
            marca.nombre = data.nombre

        if data.descripcion is not None:
            marca.descripcion = data.descripcion

        if data.estado is not None:
            marca.estado = data.estado

        return self.repository.update(
            db,
            marca,
        )

    def delete(
        self,
        db: Session,
        marca_id: int,
    ) -> None:

        marca = self.repository.get_by_id(
            db,
            marca_id,
        )
        if not marca:
            marca = self.repository.get_by_id_papelera(db, marca_id if 'marca_id' in locals() else id)

        if not marca:
            raise MarcaNoEncontradaException(MARCA_NO_EXISTE)

        if locals().get('item') and getattr(locals()['item'], 'estado', False) or (locals().get('marca') and getattr(locals().get('marca'), 'estado', False)):
            raise RegistroActivoNoPuedeEliminarseException('No se puede eliminar físicamente un registro activo. Envíelo a la papelera primero.')
        if marca.estado == True:
            raise RegistroActivoNoPuedeEliminarseException('No se puede eliminar un registro activo.')

        self.repository.delete(
            db,
            marca,
        )
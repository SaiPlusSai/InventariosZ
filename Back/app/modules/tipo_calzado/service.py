from io import BytesIO
from openpyxl import Workbook, load_workbook
from fastapi import UploadFile
from app.modules.tipo_calzado.schemas import PreviaImportacionResponse, FilaPrevia, TipoCalzadoCreate
from app.core.exceptions import RegistroActivoNoPuedeEliminarseException
from app.core.excel_utils import export_generic_excel, export_plantilla_excel
from sqlalchemy.orm import Session

from app.modules.tipo_calzado.constants import (
    TIPO_CALZADO_NO_EXISTE,
    TIPO_CALZADO_YA_EXISTE,
)
from app.modules.tipo_calzado.exceptions import (
    TipoCalzadoNoEncontradoException,
)
from app.modules.tipo_calzado.models import TipoCalzado
from app.modules.tipo_calzado.repository import TipoCalzadoRepository
from app.modules.tipo_calzado.schemas import (
    TipoCalzadoCreate,
    TipoCalzadoUpdate,
)


class TipoCalzadoService:

    def __init__(self):
        self.repository = TipoCalzadoRepository()

    
    def get_papelera(self, db: Session) -> list[TipoCalzado]:
        return self.repository.get_papelera(db)

    def get_dependencias(self, db: Session, id: int) -> dict:
        return self.repository.get_dependencias(db, id)

    def desactivar(self, db: Session, id: int):
        item = self.repository.get_by_id(db, id)
        if item:
            item.estado = False
            from datetime import datetime
            item.deleted_at = datetime.now()
            db.commit()
            db.refresh(item)
        return item

    def recuperar(self, db: Session, id: int):
        item = self.repository.get_by_id_papelera(db, id)
        if item:
            conflicto = self.repository.get_by_nombre(db, item.nombre)
            if conflicto and conflicto.id != item.id:
                from app.core.exceptions import RecuperacionConflictivaException
                raise RecuperacionConflictivaException(
                    f"No se puede recuperar. Ya existe un tipo de calzado activo con el nombre '{item.nombre}'."
                )

            item.estado = True
            item.deleted_at = None
            db.commit()
            db.refresh(item)
        return item

    
    def exportar_excel(self, db: Session) -> BytesIO:
        items = self.repository.get_all(db)
        data = [
            [
                item.id,
                item.nombre,
                item.descripcion or '',
                "Activo" if item.estado else "Inactivo"
            ]
            for item in items
        ]
        return export_generic_excel("TipoCalzado", ["ID", "Nombre", "Descripción", "Estado"], data)

    def generar_plantilla_importacion(self) -> BytesIO:
        return export_plantilla_excel(
            "Plantilla TipoCalzado", 
            ["Nombre", "Descripción"], 
            [["Deportivo", "Zapatos para hacer deporte"]]
        )

    async def previa_importacion(self, db: Session, file: UploadFile) -> PreviaImportacionResponse:
        content = await file.read()
        wb = load_workbook(filename=BytesIO(content), data_only=True)
        ws = wb.active

        filas = []
        validos = 0
        errores = 0

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue
            
            nombre, descripcion = (list(row)[:2] + [None]*2)[:2]
            
            fila_errores = []
            nombre_str = str(nombre).strip() if nombre is not None else None
            desc_str = str(descripcion).strip() if descripcion is not None else None
            

            if not nombre_str: 
                fila_errores.append("Nombre es obligatorio.")
            else:
                existente = self.repository.get_by_nombre(db, nombre_str)
                if existente:
                    fila_errores.append(f"TipoCalzado '{nombre_str}' ya existe.")
            
            

            es_valido = len(fila_errores) == 0
            if es_valido:
                validos += 1
            else:
                errores += 1

            filas.append(FilaPrevia(
                fila=idx,
                nombre=nombre_str,
                descripcion=desc_str,
                
                valido=es_valido,
                errores=fila_errores
            ))

        return PreviaImportacionResponse(total=len(filas), validos=validos, errores=errores, filas=filas)

    def confirmar_importacion(self, db: Session, filas: list[FilaPrevia]):
        try:
            creados = 0
            for fila in filas:
                if not fila.valido: continue
                payload = TipoCalzadoCreate(nombre=fila.nombre, descripcion=fila.descripcion or '')
                self.create(db, payload)
                creados += 1
            db.commit()
            return {"success": True, "creados": creados}
        except Exception as e:
            db.rollback()
            raise e

    def get_all(
        self,
        db: Session,
    ) -> list[TipoCalzado]:

        return self.repository.get_all(db)

    def get_by_id(
        self,
        db: Session,
        tipo_calzado_id: int,
    ) -> TipoCalzado | None:

        return self.repository.get_by_id(db, tipo_calzado_id)

    def create(
        self,
        db: Session,
        data: TipoCalzadoCreate,
    ) -> TipoCalzado:

        tipo_existente = self.repository.get_by_nombre_any_state(
            db,
            data.nombre,
        )

        if tipo_existente:
            if tipo_existente.estado:
                from app.core.exceptions import RegistroYaExisteException
                raise RegistroYaExisteException("El tipo de calzado ya existe.")
            else:
                from app.core.exceptions import RegistroEnPapeleraException
                raise RegistroEnPapeleraException(
                    message=f"El tipo de calzado '{data.nombre}' existe en la Papelera.",
                    id_registro=tipo_existente.id,
                    tipo_registro="tipo_calzado"
                )

        nuevo_tipo_calzado = TipoCalzado(
            nombre=data.nombre,
            descripcion=data.descripcion,
        )

        return self.repository.create(
            db,
            nuevo_tipo_calzado,
        )

    def update(
        self,
        db: Session,
        tipo_calzado_id: int,
        data: TipoCalzadoUpdate,
    ) -> TipoCalzado:

        tipo_calzado = self.repository.get_by_id(
            db,
            tipo_calzado_id,
        )

        if not tipo_calzado:
            raise TipoCalzadoNoEncontradoException(TIPO_CALZADO_NO_EXISTE)

        if data.nombre is not None:
            tipo_calzado.nombre = data.nombre

        if data.descripcion is not None:
            tipo_calzado.descripcion = data.descripcion

        if data.estado is not None:
            tipo_calzado.estado = data.estado

        return self.repository.update(
            db,
            tipo_calzado,
        )

    def delete(
        self,
        db: Session,
        tipo_calzado_id: int,
    ) -> None:

        tipo_calzado = self.repository.get_by_id(
            db,
            tipo_calzado_id,
        )
        if not tipo_calzado:
            tipo_calzado = self.repository.get_by_id_papelera(db, tipo_calzado_id if 'tipo_calzado_id' in locals() else id)

        if not tipo_calzado:
            raise TipoCalzadoNoEncontradoException(
                TIPO_CALZADO_NO_EXISTE
            )

        if locals().get('item') and getattr(locals()['item'], 'estado', False) or (locals().get('tipo_calzado') and getattr(locals().get('tipo_calzado'), 'estado', False)):
            raise RegistroActivoNoPuedeEliminarseException('No se puede eliminar físicamente un registro activo. Envíelo a la papelera primero.')
        if tipo_calzado.estado == True:
            raise RegistroActivoNoPuedeEliminarseException('No se puede eliminar un registro activo.')

        self.repository.delete(
            db,
            tipo_calzado,
        )

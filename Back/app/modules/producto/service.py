from app.core.exceptions import RegistroActivoNoPuedeEliminarseException, ValidacionDatosException
from datetime import datetime
from io import BytesIO

from sqlalchemy.orm import Session, selectinload, joinedload
from sqlalchemy import func
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from fastapi import UploadFile
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from app.modules.producto.websocket import manager
from app.modules.codigo_producto.constants import (
    CODIGO_PRODUCTO_NO_EXISTE,
    CODIGO_PRODUCTO_YA_EXISTE,
)
from app.modules.codigo_producto.exceptions import (
    CodigoProductoNoEncontradoException,
    CodigoProductoYaExisteException,
)
from app.modules.codigo_producto.models import CodigoProducto
from app.modules.codigo_producto.repository import CodigoProductoRepository
from app.modules.color.constants import COLOR_NO_EXISTE
from app.modules.color.exceptions import ColorNoEncontradoException
from app.modules.color.models import Color
from app.modules.color.repository import ColorRepository
from app.modules.marca.constants import MARCA_NO_EXISTE
from app.modules.marca.exceptions import MarcaNoEncontradaException
from app.modules.marca.repository import MarcaRepository
from app.modules.material.constants import MATERIAL_NO_EXISTE
from app.modules.material.exceptions import MaterialNoEncontradoException
from app.modules.material.repository import MaterialRepository
from app.modules.precio_producto.models import PrecioProducto
from app.modules.producto.constants import (
    PRODUCTO_NO_EXISTE,
    PRODUCTO_YA_EXISTE,
)
from app.modules.producto.exceptions import (
    ProductoNoEncontradoException,
    ProductoYaExisteException,
    StockInsuficienteException
)
from app.modules.producto.models import Producto
from app.modules.producto.repository import ProductoRepository
from app.modules.producto.schemas import (
    ColorInfo,
    MarcaInfo,
    MaterialInfo,
    ProductoCompletoCreate,
    ProductoCreate,
    ProductoListadoResponse,
    ProductoCatalogoResponse,
    ColorCatalogoResponse,
    VarianteCatalogoResponse,
    ProductoColorUpdate,
    VarianteColorUpdate,
    ProductoUpdate,
    TallaInfo,
    TipoCalzadoInfo,
    ProductoDetalleResponse,
    PrecioDetalleResponse,
    ImagenDetalleResponse,
    ProductoCompletoEditarResponse,
    ProductoCompletoUpdate,
    VarianteEditarResponse,
    ImagenEditarResponse,
    VarianteDetalleResponse,
    PreviaImportacionResponse,
    FilaPrevia,
)
from app.modules.producto_imagen.models import ProductoImagen
from app.modules.talla.constants import TALLA_NO_EXISTE
from app.modules.talla.exceptions import TallaNoEncontradaException
from app.modules.talla.models import Talla
from app.modules.talla.repository import TallaRepository
from app.modules.tipo_calzado.constants import TIPO_CALZADO_NO_EXISTE
from app.modules.tipo_calzado.exceptions import TipoCalzadoNoEncontradoException
from app.modules.tipo_calzado.repository import TipoCalzadoRepository


class ProductoService:

    def __init__(self):
        self.repository = ProductoRepository()
        self.codigo_repository = CodigoProductoRepository()
        self.marca_repository = MarcaRepository()
        self.tipo_repository = TipoCalzadoRepository()
        self.material_repository = MaterialRepository()
        self.color_repository = ColorRepository()
        self.talla_repository = TallaRepository()

    def _validar_variantes(self, variantes):
        for idx, v in enumerate(variantes):
            if v.stock_actual < 0:
                raise ValidacionDatosException(f"El stock actual en la fila {idx+1} no puede ser negativo.")
            if v.stock_minimo < 0:
                raise ValidacionDatosException(f"El stock mínimo en la fila {idx+1} no puede ser negativo.")
            if getattr(v, "precio_compra", None) is not None and v.precio_compra < 0:
                raise ValidacionDatosException(f"El precio de compra en la fila {idx+1} no puede ser negativo.")
            if getattr(v, "precio_venta", 1) <= 0:
                raise ValidacionDatosException(f"El precio de venta en la fila {idx+1} debe ser mayor a 0.")

    def get_catalogo(
        self,
        db: Session,
        codigo: str | None = None,
        marca_id: int | None = None,
        marca: str | None = None,
        color_id: int | None = None,
        color: str | None = None,
        material_id: int | None = None,
        material: str | None = None,
        talla_id: int | None = None,
        talla: str | None = None,
        tipo_calzado_id: int | None = None,
        tipo: str | None = None,
    ) -> list[ProductoCatalogoResponse]:
        productos_planos = self.repository.get_all(
            db,
            codigo=codigo,
            marca_id=marca_id,
            marca=marca,
            color_id=color_id,
            color=color,
            material_id=material_id,
            material=material,
            talla_id=talla_id,
            talla=talla,
            tipo_calzado_id=tipo_calzado_id,
            tipo=tipo,
        )

        catalogo_dict = {}

        for p in productos_planos:
            agrupador = (p["codigo_producto_id"], p["tipo_calzado_id"], p["material_id"], p["descripcion"])
            
            if agrupador not in catalogo_dict:
                catalogo_dict[agrupador] = {
                    "codigo_producto_id": p["codigo_producto_id"],
                    "producto_principal_id": p["id"],
                    "codigo": p["codigo"],
                    "marca": MarcaInfo(id=p["marca_id"], nombre=p["marca_nombre"]),
                    "tipo_calzado": TipoCalzadoInfo(id=p["tipo_calzado_id"], nombre=p["tipo_calzado_nombre"]),
                    "material": MaterialInfo(id=p["material_id"], nombre=p["material_nombre"]),
                    "descripcion": p["descripcion"],
                    "colores": {}
                }

            colores_dict = catalogo_dict[agrupador]["colores"]
            color_id = p["color_id"]

            if color_id not in colores_dict:
                colores_dict[color_id] = {
                    "color_id": color_id,
                    "color": ColorInfo(id=color_id, nombre=p["color_nombre"]),
                    "imagen_principal": p["imagen_principal"],
                    "variantes": []
                }
            elif p["imagen_principal"] and not colores_dict[color_id]["imagen_principal"]:
                colores_dict[color_id]["imagen_principal"] = p["imagen_principal"]

            variante = VarianteCatalogoResponse(
                id=p["id"],
                talla=TallaInfo(id=p["talla_id"], nombre=p["talla_nombre"]),
                stock_actual=p["stock_actual"],
                stock_minimo=p["stock_minimo"],
                stock_maximo=p["stock_maximo"],
                precio_compra=p["precio_compra"],
                precio_venta=p["precio_venta"],
                estado=p["estado"]
            )
            colores_dict[color_id]["variantes"].append(variante)

        respuestas = []
        for cp_id, cp_data in catalogo_dict.items():
            colores_list = [
                ColorCatalogoResponse(
                    color_id=c_data["color_id"],
                    color=c_data["color"],
                    imagen_principal=c_data["imagen_principal"],
                    variantes=c_data["variantes"]
                )
                for c_data in cp_data["colores"].values()
            ]
            respuestas.append(
                ProductoCatalogoResponse(
                    codigo_producto_id=cp_data["codigo_producto_id"],
                    producto_principal_id=cp_data["producto_principal_id"],
                    codigo=cp_data["codigo"],
                    marca=cp_data["marca"],
                    tipo_calzado=cp_data["tipo_calzado"],
                    material=cp_data["material"],
                    descripcion=cp_data["descripcion"],
                    colores=colores_list
                )
            )

        return respuestas

    def exportar_excel(
        self,
        db: Session,
        codigo: str | None = None,
        marca_id: int | None = None,
        marca: str | None = None,
        color_id: int | None = None,
        color: str | None = None,
        material_id: int | None = None,
        material: str | None = None,
        talla_id: int | None = None,
        talla: str | None = None,
        tipo_calzado_id: int | None = None,
        tipo: str | None = None,
    ) -> BytesIO:
        productos_planos = self.repository.get_all(
            db,
            codigo=codigo,
            marca_id=marca_id,
            marca=marca,
            color_id=color_id,
            color=color,
            material_id=material_id,
            material=material,
            talla_id=talla_id,
            talla=talla,
            tipo_calzado_id=tipo_calzado_id,
            tipo=tipo,
        )

        from app.core.excel_utils import export_generic_excel

        data = []
        for p in productos_planos:
            data.append([
                p["codigo"],
                p["marca_nombre"],
                p["tipo_calzado_nombre"],
                p["material_nombre"],
                p["color_nombre"],
                p["talla_nombre"],
                int(p["stock_actual"]),
                float(p["precio_venta"]),
                "Activo" if p["estado"] else "Inactivo"
            ])
            
        return export_generic_excel(
            "Productos", 
            ["Código", "Marca", "Tipo", "Material", "Color", "Talla", "Stock", "Precio Venta", "Estado"], 
            data
        )

    def exportar_pdf(
        self,
        db: Session,
        codigo: str | None = None,
        marca_id: int | None = None,
        marca: str | None = None,
        color_id: int | None = None,
        color: str | None = None,
        material_id: int | None = None,
        material: str | None = None,
        talla_id: int | None = None,
        talla: str | None = None,
        tipo_calzado_id: int | None = None,
        tipo: str | None = None,
    ) -> BytesIO:
        productos_planos = self.repository.get_all(
            db,
            codigo=codigo,
            marca_id=marca_id,
            marca=marca,
            color_id=color_id,
            color=color,
            material_id=material_id,
            material=material,
            talla_id=talla_id,
            talla=talla,
            tipo_calzado_id=tipo_calzado_id,
            tipo=tipo,
        )

        buffer = BytesIO()
        # Use margins to make it look cleaner
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                                rightMargin=40, leftMargin=40,
                                topMargin=40, bottomMargin=50)
        elements = []
        styles = getSampleStyleSheet()

        # Custom Styles
        title_style = ParagraphStyle(
            'CustomTitle', 
            parent=styles['Heading1'], 
            fontSize=22, 
            textColor=colors.HexColor("#2C3E50"),
            spaceAfter=6,
            alignment=1 # Center
        )
        subtitle_style = ParagraphStyle(
            'CustomSubtitle', 
            parent=styles['Normal'], 
            fontSize=12, 
            textColor=colors.HexColor("#7F8C8D"),
            spaceAfter=20,
            alignment=1 # Center
        )

        nombre_negocio = "MI NEGOCIO S.A."
        titulo_reporte = "Catálogo General de Inventario"

        elements.append(Paragraph(f"<b>{nombre_negocio}</b>", title_style))
        elements.append(Paragraph(f"{titulo_reporte} — Total de productos: {len(productos_planos)}", subtitle_style))
        elements.append(Spacer(1, 0.2 * inch))

        # Tabla
        data = [[
            "Código", "Marca", "Tipo", "Material", "Color", 
            "Talla", "Stock", "Precio Venta", "Estado"
        ]]

        for p in productos_planos:
            data.append([
                p["codigo"],
                p["marca_nombre"],
                p["tipo_calzado_nombre"],
                p["material_nombre"],
                p["color_nombre"],
                p["talla_nombre"],
                str(p["stock_actual"]),
                f"${p['precio_venta']:.2f}",
                "Activo" if p["estado"] else "Inactivo"
            ])

        # Alternating row colors
        row_colors = []
        for i in range(1, len(data)):
            if i % 2 == 0:
                row_colors.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor("#F8F9F9")))
            else:
                row_colors.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor("#FFFFFF")))

        base_style = [
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2980B9")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            
            # Body
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            
            # Borders
            ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor("#BDC3C7")),
        ]

        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle(base_style + row_colors))
        elements.append(t)

        def add_footer(canvas, doc):
            canvas.saveState()
            # Línea separadora
            canvas.setStrokeColor(colors.HexColor("#BDC3C7"))
            canvas.setLineWidth(1)
            canvas.line(40, 40, landscape(letter)[0] - 40, 40)
            
            # Fecha de descarga
            fecha_str = datetime.now().strftime("%d de %B, %Y a las %H:%M")
            canvas.setFont('Helvetica-Oblique', 9)
            canvas.setFillColor(colors.HexColor("#7F8C8D"))
            canvas.drawString(40, 25, f"Generado el: {fecha_str}")
            
            # Paginación
            page_num = canvas.getPageNumber()
            text = f"Página {page_num}"
            canvas.drawRightString(landscape(letter)[0] - 40, 25, text)
            
            canvas.restoreState()

        doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)
        buffer.seek(0)
        return buffer

    def generar_plantilla_importacion(self) -> BytesIO:
        from app.core.excel_utils import export_plantilla_excel
        return export_plantilla_excel(
            "Plantilla Importacion Productos",
            [
                "Código", "Marca", "Tipo", "Material", "Color", 
                "Talla", "Stock", "Precio Venta", "Descripción"
            ],
            [[
                "PROD-001", "Nike", "Urbano", "Cuero", "Blanco", 
                "40", 100, 150.00, "Zapatilla de prueba"
            ]]
        )

    async def previa_importacion(self, db: Session, file: UploadFile) -> PreviaImportacionResponse:
        content = await file.read()
        wb = load_workbook(filename=BytesIO(content), data_only=True)
        ws = wb.active

        filas = []
        validos = 0
        errores = 0

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue
            
            codigo, marca, tipo, material, color, talla, stock, precio, descripcion = (list(row)[:9] + [None]*9)[:9]

            fila_errores = []

            codigo_str = str(codigo).strip() if codigo is not None else None
            marca_str = str(marca).strip() if marca is not None else None
            tipo_str = str(tipo).strip() if tipo is not None else None
            material_str = str(material).strip() if material is not None else None
            color_str = str(color).strip() if color is not None else None
            talla_str = str(talla).strip() if talla is not None else None
            stock_str = str(stock).strip() if stock is not None else None
            precio_str = str(precio).strip() if precio is not None else None
            descripcion_str = str(descripcion).strip() if descripcion is not None else None

            if not codigo_str: fila_errores.append("Código es obligatorio.")
            else:
                existente = self.codigo_repository.get_by_codigo(db, codigo_str)
                if existente:
                    fila_errores.append(f"Código '{codigo_str}' ya existe en BD.")
            
            if not marca_str: fila_errores.append("Marca es obligatoria.")
            else:
                if not self.marca_repository.get_by_nombre(db, marca_str):
                    fila_errores.append(f"Marca '{marca_str}' no existe.")

            if not tipo_str: fila_errores.append("Tipo es obligatorio.")
            else:
                if not self.tipo_repository.get_by_nombre(db, tipo_str):
                    fila_errores.append(f"Tipo '{tipo_str}' no existe.")

            if not material_str: fila_errores.append("Material es obligatorio.")
            else:
                if not self.material_repository.get_by_nombre(db, material_str):
                    fila_errores.append(f"Material '{material_str}' no existe.")

            if not color_str: fila_errores.append("Color es obligatorio.")
            else:
                if not self.color_repository.get_by_nombre(db, color_str):
                    fila_errores.append(f"Color '{color_str}' no existe.")

            if not talla_str: fila_errores.append("Talla es obligatoria.")
            else:
                if not self.talla_repository.get_by_nombre(db, talla_str):
                    fila_errores.append(f"Talla '{talla_str}' no existe.")

            if not stock_str: fila_errores.append("Stock es obligatorio.")
            else:
                try:
                    s = int(float(stock_str))
                    if s < 0: fila_errores.append("Stock debe ser >= 0.")
                except ValueError:
                    fila_errores.append("Stock debe ser un número entero.")

            if not precio_str: fila_errores.append("Precio Venta es obligatorio.")
            else:
                try:
                    p = float(precio_str)
                    if p < 0: fila_errores.append("Precio debe ser >= 0.")
                except ValueError:
                    fila_errores.append("Precio debe ser un número válido.")
            
            es_valido = len(fila_errores) == 0
            if es_valido:
                validos += 1
            else:
                errores += 1

            filas.append(FilaPrevia(
                fila=idx,
                codigo=codigo_str,
                marca=marca_str,
                tipo=tipo_str,
                material=material_str,
                color=color_str,
                talla=talla_str,
                stock=stock_str,
                precio=precio_str,
                descripcion=descripcion_str,
                valido=es_valido,
                errores=fila_errores
            ))
            
        return PreviaImportacionResponse(
            total=len(filas),
            validos=validos,
            errores=errores,
            filas=filas
        )

    def confirmar_importacion(self, db: Session, filas: list[FilaPrevia]):
        productos_a_crear = {}
        for fila in filas:
            if not fila.valido: continue
            
            marca = self.marca_repository.get_by_nombre(db, fila.marca)
            tipo = self.tipo_repository.get_by_nombre(db, fila.tipo)
            material = self.material_repository.get_by_nombre(db, fila.material)
            color = self.color_repository.get_by_nombre(db, fila.color)
            talla = self.talla_repository.get_by_nombre(db, fila.talla)
            
            codigo = fila.codigo
            if codigo not in productos_a_crear:
                productos_a_crear[codigo] = {
                    "marca_id": marca.id,
                    "tipo_calzado_id": tipo.id,
                    "material_id": material.id,
                    "descripcion": fila.descripcion or "",
                    "variantes": []
                }
            
            productos_a_crear[codigo]["variantes"].append(
                VarianteCreate(
                    color_id=color.id,
                    talla_id=talla.id,
                    stock_actual=int(float(fila.stock)),
                    precio_venta=float(fila.precio)
                )
            )

        # Envolver en transacción
        try:
            creados = 0
            for codigo, props in productos_a_crear.items():
                payload = ProductoCompletoCreate(
                    codigo=codigo,
                    marca_id=props["marca_id"],
                    tipo_calzado_id=props["tipo_calzado_id"],
                    material_id=props["material_id"],
                    descripcion=props["descripcion"],
                    variantes=props["variantes"],
                    force=True
                )
                self.create_completo(db, payload)
                creados += 1
            db.commit()
            return {"success": True, "creados": creados}
        except Exception as e:
            db.rollback()
            raise e

    def get_papelera(
        self,
        db: Session,
        codigo: str | None = None,
        marca_id: int | None = None,
        marca: str | None = None,
        color_id: int | None = None,
        color: str | None = None,
        material_id: int | None = None,
        material: str | None = None,
        talla_id: int | None = None,
        talla: str | None = None,
        tipo_calzado_id: int | None = None,
        tipo: str | None = None,
    ) -> list[ProductoListadoResponse]:
        productos = self.repository.get_papelera(
            db,
            codigo=codigo,
            marca_id=marca_id,
            marca=marca,
            color_id=color_id,
            color=color,
            material_id=material_id,
            material=material,
            talla_id=talla_id,
            talla=talla,
            tipo_calzado_id=tipo_calzado_id,
            tipo=tipo,
        )

        return [
            ProductoListadoResponse(
                id=producto["id"],
                codigo_producto_id=producto["codigo_producto_id"],
                codigo=producto["codigo"],
                descripcion=producto["descripcion"],
                marca=MarcaInfo(
                    id=producto["marca_id"],
                    nombre=producto["marca_nombre"],
                ),
                tipo_calzado=TipoCalzadoInfo(
                    id=producto["tipo_calzado_id"],
                    nombre=producto["tipo_calzado_nombre"],
                ),
                material=MaterialInfo(
                    id=producto["material_id"],
                    nombre=producto["material_nombre"],
                ),
                color=ColorInfo(
                    id=producto["color_id"],
                    nombre=producto["color_nombre"],
                ),
                talla=TallaInfo(
                    id=producto["talla_id"],
                    nombre=producto["talla_nombre"],
                ),
                stock_actual=producto["stock_actual"],
                stock_minimo=producto["stock_minimo"],
                stock_maximo=producto["stock_maximo"],
                precio_compra=producto["precio_compra"],
                precio_venta=producto["precio_venta"],
                imagen_principal=producto["imagen_principal"],
                estado=producto["estado"],
                created_at=producto["created_at"],
                updated_at=producto["updated_at"],
            )
            for producto in productos
        ]

    def get_dependencias(self, db: Session, id: int) -> dict:
        return self.repository.get_dependencias(db, id)

    def desactivar(self, db: Session, id: int):
        item = self.repository.get_by_id(db, id)
        if item:
            item.estado = False
            item.deleted_at = datetime.now()
            db.commit()
            db.refresh(item)
        return item

    def recuperar(self, db: Session, id: int):
        item = self.repository.get_by_id_papelera(db, id)
        if item:
            item.estado = True
            item.deleted_at = None
            db.commit()
            db.refresh(item)
        return item

    def get_all(
        self,
        db: Session,
        codigo: str | None = None,
        marca_id: int | None = None,
        marca: str | None = None,
        color_id: int | None = None,
        color: str | None = None,
        material_id: int | None = None,
        material: str | None = None,
        talla_id: int | None = None,
        talla: str | None = None,
        tipo_calzado_id: int | None = None,
        tipo: str | None = None,
    ) -> list[ProductoListadoResponse]:

        productos = self.repository.get_all(
            db,
            codigo=codigo,
            marca_id=marca_id,
            marca=marca,
            color_id=color_id,
            color=color,
            material_id=material_id,
            material=material,
            talla_id=talla_id,
            talla=talla,
            tipo_calzado_id=tipo_calzado_id,
            tipo=tipo,
        )

        return [
            ProductoListadoResponse(
                id=producto["id"],
                codigo_producto_id=producto["codigo_producto_id"],
                codigo=producto["codigo"],
                descripcion=producto["descripcion"],
                marca=MarcaInfo(
                    id=producto["marca_id"],
                    nombre=producto["marca_nombre"],
                ),
                tipo_calzado=TipoCalzadoInfo(
                    id=producto["tipo_calzado_id"],
                    nombre=producto["tipo_calzado_nombre"],
                ),
                material=MaterialInfo(
                    id=producto["material_id"],
                    nombre=producto["material_nombre"],
                ),
                color=ColorInfo(
                    id=producto["color_id"],
                    nombre=producto["color_nombre"],
                ),
                talla=TallaInfo(
                    id=producto["talla_id"],
                    nombre=producto["talla_nombre"],
                ),
                stock_actual=producto["stock_actual"],
                stock_minimo=producto["stock_minimo"],
                stock_maximo=producto["stock_maximo"],
                precio_compra=producto["precio_compra"],
                precio_venta=producto["precio_venta"],
                imagen_principal=producto["imagen_principal"],
                estado=producto["estado"],
                created_at=producto["created_at"],
                updated_at=producto["updated_at"],
            )
            for producto in productos
        ]
    def get_detalle(
        self,
        db: Session,
        producto_id: int,
    ) -> ProductoDetalleResponse:

        producto = self.repository.get_detalle(
            db,
            producto_id,
        )

        if not producto:
            producto = self.repository.get_by_id_papelera(db, producto_id)
        if not producto:
            raise ProductoNoEncontradoException(
                PRODUCTO_NO_EXISTE
            )

        precio = next(
            (
                p
                for p in producto.precios
                if p.estado
            ),
            None,
        )

        todas_variantes = self.repository.get_by_codigo_producto_id(
            db,
            producto.codigo_producto_id,
        )

        variantes_response = []
        for v in todas_variantes:
            v_precio = next((p for p in v.precios if p.estado), None)
            variantes_response.append(
                VarianteDetalleResponse(
                    id=v.id,
                    color=ColorInfo(
                        id=v.color.id,
                        nombre=v.color.nombre,
                        codigo_hex=v.color.codigo_hex,
                    ),
                    talla=TallaInfo(
                        id=v.talla.id,
                        nombre=v.talla.nombre,
                    ),
                    stock_actual=v.stock_actual,
                    stock_minimo=v.stock_minimo,
                    stock_maximo=v.stock_maximo,
                    precio=(
                        PrecioDetalleResponse(
                            precio_compra=v_precio.precio_compra,
                            precio_venta=v_precio.precio_venta,
                        )
                        if v_precio else None
                    ),
                    estado=v.estado,
                )
            )

        imagenes_a_mostrar = producto.imagenes

        if not imagenes_a_mostrar:
            variante_con_imagen = next(
                (v for v in todas_variantes if v.color_id == producto.color_id and v.imagenes),
                None
            )
            if variante_con_imagen:
                imagenes_a_mostrar = variante_con_imagen.imagenes

        imagen_principal = next(
            (
                img.ruta
                for img in imagenes_a_mostrar
                if img.es_principal
            ),
            None,
        )

        return ProductoDetalleResponse(

            id=producto.id,

            codigo=producto.codigo_producto.codigo,

            descripcion=producto.descripcion,

            marca=MarcaInfo(
                id=producto.codigo_producto.marca.id,
                nombre=producto.codigo_producto.marca.nombre,
            ),

            tipo_calzado=TipoCalzadoInfo(
                id=producto.tipo_calzado.id,
                nombre=producto.tipo_calzado.nombre,
            ),

            material=MaterialInfo(
                id=producto.material.id,
                nombre=producto.material.nombre,
            ),

            color=ColorInfo(
                id=producto.color.id,
                nombre=producto.color.nombre,
                codigo_hex=producto.color.codigo_hex,
            ),

            talla=TallaInfo(
                id=producto.talla.id,
                nombre=producto.talla.nombre,
            ),

            stock_actual=producto.stock_actual,
            stock_minimo=producto.stock_minimo,
            stock_maximo=producto.stock_maximo,

            precio=(
                PrecioDetalleResponse(
                    precio_compra=precio.precio_compra,
                    precio_venta=precio.precio_venta,
                )
                if precio
                else None
            ),

            imagen_principal=imagen_principal,

            imagenes=[
                ImagenDetalleResponse(
                    id=img.id,
                    bucket=img.bucket,
                    ruta=img.ruta,
                    nombre_archivo=img.nombre_archivo,
                    es_principal=img.es_principal,
                    orden=img.orden,
                )
                for img in imagenes_a_mostrar
            ],
            
            variantes=variantes_response,

            estado=producto.estado,

            created_at=producto.created_at,

            updated_at=producto.updated_at,
    )    
    def get_by_id(
        self,
        db: Session,
        producto_id: int,
    ) -> Producto | None:

        return self.repository.get_by_id(
            db,
            producto_id,
        )

    def create(
        self,
        db: Session,
        data: ProductoCreate,
    ) -> Producto:

        if not self.codigo_repository.get_by_id(db, data.codigo_producto_id):
            raise CodigoProductoNoEncontradoException(
                CODIGO_PRODUCTO_NO_EXISTE
            )

        if not self.tipo_repository.get_by_id(db, data.tipo_calzado_id):
            raise TipoCalzadoNoEncontradoException(
                TIPO_CALZADO_NO_EXISTE
            )

        if not self.material_repository.get_by_id(db, data.material_id):
            raise MaterialNoEncontradoException(
                MATERIAL_NO_EXISTE
            )

        if not self.color_repository.get_by_id(db, data.color_id):
            raise ColorNoEncontradoException(
                COLOR_NO_EXISTE
            )

        if not self.talla_repository.get_by_id(db, data.talla_id):
            raise TallaNoEncontradaException(
                TALLA_NO_EXISTE
            )

        if self.repository.exists(
            db,
            data.codigo_producto_id,
            data.color_id,
            data.talla_id,
        ):
            raise ProductoYaExisteException(
                PRODUCTO_YA_EXISTE
            )

        producto = Producto(**data.model_dump())

        return self.repository.create(
            db,
            producto,
        )
    def get_editar_completo(
        self,
        db: Session,
        grupo_id: int,
    ) -> ProductoCompletoEditarResponse:

        primer_producto = db.query(Producto).options(
            selectinload(Producto.imagenes),
            joinedload(Producto.codigo_producto)
        ).filter(Producto.id == grupo_id).first()
        if not primer_producto:
            raise ProductoNoEncontradoException(PRODUCTO_NO_EXISTE)

        productos = db.query(Producto).options(
            selectinload(Producto.precios)
        ).filter(
            Producto.codigo_producto_id == primer_producto.codigo_producto_id,
            Producto.tipo_calzado_id == primer_producto.tipo_calzado_id,
            Producto.material_id == primer_producto.material_id,
            Producto.descripcion == primer_producto.descripcion
        ).all()

        variantes = []

        for producto in productos:

            precio = next(
                (
                    p
                    for p in producto.precios
                    if p.estado
                ),
                None,
            )

            variantes.append(
                VarianteEditarResponse(
                    id=producto.id,
                    color_id=producto.color_id,
                    talla_id=producto.talla_id,
                    stock_actual=producto.stock_actual,
                    stock_minimo=producto.stock_minimo,
                    stock_maximo=producto.stock_maximo,
                    precio_compra=(
                        precio.precio_compra
                        if precio
                        else None
                    ),
                    precio_venta=(
                        precio.precio_venta
                        if precio
                        else 0
                    ),
                )
            )

        imagenes = [
            ImagenEditarResponse(
                id=imagen.id,
                bucket=imagen.bucket,
                ruta=imagen.ruta,
                nombre_archivo=imagen.nombre_archivo,
                es_principal=imagen.es_principal,
                orden=imagen.orden,
            )
            for imagen in primer_producto.imagenes
        ]

        return ProductoCompletoEditarResponse(
            codigo_producto_id=primer_producto.codigo_producto.id,
            codigo=primer_producto.codigo_producto.codigo,
            marca_id=primer_producto.codigo_producto.marca_id,
            tipo_calzado_id=primer_producto.tipo_calzado_id,
            material_id=primer_producto.material_id,
            descripcion=primer_producto.descripcion,
            variantes=variantes,
            imagenes=imagenes,
        )
    def update_completo(
        self,
        db: Session,
        grupo_id: int,
        data: ProductoCompletoUpdate,
    ):
        self._validar_variantes(data.variantes)
        """
        Edita completamente un grupo lógico de productos.
        """

        try:
            representante = self.repository.get_by_id(db, grupo_id)
            if not representante:
                raise ProductoNoEncontradoException(PRODUCTO_NO_EXISTE)

            codigo_producto = self.codigo_repository.get_by_id(
                db,
                data.codigo_producto_id,
            )

            if not codigo_producto:
                raise ProductoNoEncontradoException(
                    PRODUCTO_NO_EXISTE
                )



            if not self.tipo_repository.get_by_id(
                db,
                data.tipo_calzado_id,
            ):
                raise TipoCalzadoNoEncontradoException(
                    TIPO_CALZADO_NO_EXISTE
                )

            if not self.material_repository.get_by_id(
                db,
                data.material_id,
            ):
                raise MaterialNoEncontradoException(
                    MATERIAL_NO_EXISTE
                )

            variantes_unicas = set()
            cache_colores = {}
            cache_tallas = {}

            for variante in data.variantes:
                clave = (
                    variante.color_id,
                    variante.talla_id,
                )

                if clave in variantes_unicas:
                    raise ProductoYaExisteException(
                        PRODUCTO_YA_EXISTE
                    )

                variantes_unicas.add(clave)

                if variante.color_id not in cache_colores:
                    color_obj = self.color_repository.get_by_id(db, variante.color_id)
                    if not color_obj:
                        raise ColorNoEncontradoException(COLOR_NO_EXISTE)
                    cache_colores[variante.color_id] = color_obj

                if variante.talla_id not in cache_tallas:
                    talla_obj = self.talla_repository.get_by_id(db, variante.talla_id)
                    if not talla_obj:
                        raise TallaNoEncontradaException(TALLA_NO_EXISTE)
                    cache_tallas[variante.talla_id] = talla_obj

            # Obtenemos todos los productos actuales para este grupo logico
            productos_existentes = db.query(Producto).filter(
                Producto.codigo_producto_id == representante.codigo_producto_id,
                Producto.tipo_calzado_id == representante.tipo_calzado_id,
                Producto.material_id == representante.material_id,
                Producto.descripcion == representante.descripcion
            ).all()

            mapa_existentes = {(p.color_id, p.talla_id): p for p in productos_existentes}

            variantes_entrantes_claves = set()
            productos_modificados_o_creados = []

            for variante in data.variantes:
                clave = (variante.color_id, variante.talla_id)
                variantes_entrantes_claves.add(clave)

                if clave in mapa_existentes:
                    producto = mapa_existentes[clave]
                    # Upsert (Actualización in-place)
                    producto.tipo_calzado_id = data.tipo_calzado_id
                    producto.material_id = data.material_id
                    producto.descripcion = data.descripcion
                    producto.stock_actual = variante.stock_actual
                    producto.stock_minimo = variante.stock_minimo
                    producto.stock_maximo = variante.stock_maximo
                    producto.estado = variante.estado
                    
                    # Logica de Precio: 
                    precio_actual = None
                    for p in producto.precios:
                        if p.estado and p.vigente_hasta is None:
                            precio_actual = p
                            break
                            
                    if not precio_actual or precio_actual.precio_compra != variante.precio_compra or precio_actual.precio_venta != variante.precio_venta:
                        if precio_actual:
                            precio_actual.vigente_hasta = datetime.now()
                            precio_actual.estado = False
                            
                        nuevo_precio = PrecioProducto(
                            precio_compra=variante.precio_compra,
                            precio_venta=variante.precio_venta,
                            vigente_desde=datetime.now(),
                            estado=True,
                        )
                        producto.precios.append(nuevo_precio)
                        
                    productos_modificados_o_creados.append(producto)
                else:
                    # Insert (Nuevo)
                    # Si la variante es nueva pero viene desactivada y sin datos relevantes,
                    # asumimos que es basura de la matriz rectangular del frontend y la omitimos.
                    if (
                        not variante.estado 
                        and (variante.stock_actual == 0 or variante.stock_actual is None)
                        and (variante.precio_venta == 0 or variante.precio_venta is None)
                    ):
                        continue
                        
                    producto = Producto(
                        codigo_producto_id=codigo_producto.id,
                        tipo_calzado_id=data.tipo_calzado_id,
                        material_id=data.material_id,
                        color_id=variante.color_id,
                        talla_id=variante.talla_id,
                        descripcion=data.descripcion,
                        stock_actual=variante.stock_actual,
                        stock_minimo=variante.stock_minimo,
                        stock_maximo=variante.stock_maximo,
                        estado=variante.estado,
                    )
                    
                    nuevo_precio = PrecioProducto(
                        precio_compra=variante.precio_compra,
                        precio_venta=variante.precio_venta,
                        vigente_desde=datetime.now(),
                        estado=True,
                    )
                    producto.precios.append(nuevo_precio)
                    db.add(producto)
                    productos_modificados_o_creados.append(producto)
                    
            # Las que no vinieron en el payload, se desactivan logicamente
            for clave, producto in mapa_existentes.items():
                if clave not in variantes_entrantes_claves:
                    producto.estado = False

            db.commit()

            return {
                "codigo_producto_id": codigo_producto.id,
                "producto_principal_id": productos_modificados_o_creados[0].id if productos_modificados_o_creados else None,
                "variantes_creadas": len(productos_modificados_o_creados),
                "precios_creados": len(productos_modificados_o_creados),
                "imagenes_creadas": 0,
                "success": True,
                "message": "Producto actualizado correctamente.",
                "created_at": datetime.now(),
            }

        except Exception:

            self.repository.rollback(
                db,
            )

            raise
    def update(
        self,
        db: Session,
        producto_id: int,
        data: ProductoUpdate,
    ) -> Producto:

        producto = self.repository.get_by_id(
            db,
            producto_id,
        )

        if not producto:
            producto = self.repository.get_by_id_papelera(db, producto_id)
        if not producto:
            raise ProductoNoEncontradoException(
                PRODUCTO_NO_EXISTE
            )

        update_data = data.model_dump(
            exclude_unset=True
        )

        for campo, valor in update_data.items():
            setattr(
                producto,
                campo,
                valor,
            )

        return self.repository.update(
            db,
            producto,
        )

    def delete(
        self,
        db: Session,
        producto_id: int,
    ) -> None:

        producto = self.repository.get_by_id(
            db,
            producto_id,
        )

        if not producto:
            producto = self.repository.get_by_id_papelera(db, producto_id)
        if not producto:
            producto = self.repository.get_by_id_papelera(db, producto_id if 'producto_id' in locals() else id)
        if not producto:
            raise ProductoNoEncontradoException(
                PRODUCTO_NO_EXISTE
            )

        if locals().get('item') and getattr(locals()['item'], 'estado', False) or (locals().get('producto') and getattr(locals().get('producto'), 'estado', False)):
            raise RegistroActivoNoPuedeEliminarseException('No se puede eliminar físicamente un registro activo. Envíelo a la papelera primero.')
        if producto.estado == True:
            raise RegistroActivoNoPuedeEliminarseException(
                "No se puede eliminar un registro activo. Desactívalo primero."
            )

        self.repository.delete(db, producto)

    def create_completo(
        self,
        db: Session,
        data: ProductoCompletoCreate,
    ) -> dict:
        
        self._validar_variantes(data.variantes)
        """
        Crea un producto completo en una sola transaccion.
        """

        try:

            if not self.tipo_repository.get_by_id(
                db,
                data.tipo_calzado_id,
            ):
                raise TipoCalzadoNoEncontradoException(
                    TIPO_CALZADO_NO_EXISTE
                )

            if not self.material_repository.get_by_id(
                db,
                data.material_id,
            ):
                raise MaterialNoEncontradoException(
                    MATERIAL_NO_EXISTE
                )

            variantes_unicas = set()
            color_ids = list({v.color_id for v in data.variantes})
            talla_ids = list({v.talla_id for v in data.variantes})

            colores_db = db.query(Color).filter(Color.id.in_(color_ids)).all()
            tallas_db = db.query(Talla).filter(Talla.id.in_(talla_ids)).all()

            cache_colores = {c.id: c for c in colores_db}
            cache_tallas = {t.id: t for t in tallas_db}

            for c_id in color_ids:
                if c_id not in cache_colores:
                    raise ColorNoEncontradoException(COLOR_NO_EXISTE)
            for t_id in talla_ids:
                if t_id not in cache_tallas:
                    raise TallaNoEncontradaException(TALLA_NO_EXISTE)

            for variante in data.variantes:
                clave_variante = (variante.color_id, variante.talla_id)

                if clave_variante in variantes_unicas:
                    raise ProductoYaExisteException(PRODUCTO_YA_EXISTE)

                variantes_unicas.add(clave_variante)

            print("✔ Talla encontrada")

            codigo_producto = self.codigo_repository.get_by_id(db, data.codigo_producto_id)
            if not codigo_producto:
                raise CodigoProductoNoEncontradoException(CODIGO_PRODUCTO_NO_EXISTE)
            
            print("✔ Código encontrado")

            productos = []

            # 1. Consultar BD para detectar colisiones antes de iterar
            productos_existentes = db.query(Producto).filter(
                Producto.codigo_producto_id == data.codigo_producto_id,
                Producto.tipo_calzado_id == data.tipo_calzado_id,
                Producto.material_id == data.material_id,
                Producto.descripcion == data.descripcion
            ).all()
            
            # Guardamos el estado de cada variante existente
            mapa_existentes = {(p.color_id, p.talla_id): p.estado for p in productos_existentes}

            for variante in data.variantes:
                # 2. Filtrar variantes inactivas que son basura de la matriz rectangular
                if (
                    not variante.estado 
                    and (variante.stock_actual == 0 or variante.stock_actual is None)
                    and (variante.precio_venta == 0 or variante.precio_venta is None)
                ):
                    continue

                # 3. Validar si ya existe en la BD (sea activa o inactiva)
                clave = (variante.color_id, variante.talla_id)
                if clave in mapa_existentes:
                    if mapa_existentes[clave]:
                        raise ProductoYaExisteException(PRODUCTO_YA_EXISTE)
                    else:
                        raise ProductoYaExisteException("La variante ya existe pero está inactiva. Por favor utilice el flujo de Edición para reactivarla.")

                producto = Producto(
                    codigo_producto=codigo_producto,
                    tipo_calzado_id=data.tipo_calzado_id,
                    material_id=data.material_id,
                    color_id=variante.color_id,
                    talla_id=variante.talla_id,
                    descripcion=data.descripcion,
                    stock_actual=variante.stock_actual,
                    stock_minimo=variante.stock_minimo,
                    stock_maximo=variante.stock_maximo,
                    estado=variante.estado,
                )

                nuevo_precio = PrecioProducto(
                    precio_compra=variante.precio_compra,
                    precio_venta=variante.precio_venta,
                    vigente_desde=datetime.now(),
                    estado=True,
                )
                producto.precios.append(nuevo_precio)
                db.add(producto)
                productos.append(producto)

            print("✔ Producto insertado en sesión")
            db.flush()
            print("✔ Flush realizado")

            # Eliminado bucle db.refresh() redundante. db.flush() ya pobló los IDs autoincrementales.

            print("Producto ID:", productos[0].id if productos else None)

            print("✔ Variantes creadas")
            print("✔ Precios creados")
            print("✔ Imágenes creadas")

            db.commit()
            print("✔ Commit realizado")

            res = {
                "codigo_producto_id": codigo_producto.id,
                "producto_principal_id": productos[0].id if productos else None,
                "variantes_creadas": len(productos),
                "precios_creados": len(productos),
                "imagenes_creadas": 0,
                "success": True,
                "message": "Producto creado correctamente.",
                "created_at": datetime.now(),
            }
            print("✔ Respuesta enviada")
            return res

        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            print("=" * 80)
            print("ERROR EN crear_completo")
            print(tb)
            print(type(e))
            print(e)
            db.rollback()
            raise Exception(f"{str(e)}\n\nTraceback:\n{tb}")

    def desactivar_color(self, db: Session, grupo_id: int, color_id: int, commit: bool = True, representante: Producto = None):
        if not representante:
            representante = self.repository.get_by_id(db, grupo_id)
        if not representante:
            raise ProductoNoEncontradoException(PRODUCTO_NO_EXISTE)
        variantes = db.query(Producto).filter(
            Producto.codigo_producto_id == representante.codigo_producto_id,
            Producto.tipo_calzado_id == representante.tipo_calzado_id,
            Producto.material_id == representante.material_id,
            Producto.descripcion == representante.descripcion,
            Producto.color_id == color_id,
            Producto.estado == True
        ).all()
        for v in variantes:
            v.estado = False
            v.deleted_at = datetime.now()
        if commit:
            db.commit()
        return {'msg': f'{len(variantes)} variantes desactivadas'}

    def recuperar_color(self, db: Session, grupo_id: int, color_id: int, commit: bool = True, representante: Producto = None):
        if not representante:
            representante = self.repository.get_by_id_including_deleted(db, grupo_id)
        if not representante:
            raise ProductoNoEncontradoException(PRODUCTO_NO_EXISTE)
        variantes = db.query(Producto).filter(
            Producto.codigo_producto_id == representante.codigo_producto_id,
            Producto.tipo_calzado_id == representante.tipo_calzado_id,
            Producto.material_id == representante.material_id,
            Producto.descripcion == representante.descripcion,
            Producto.color_id == color_id,
            Producto.estado == False
        ).all()
        for v in variantes:
            v.estado = True
            v.deleted_at = None
        if commit:
            db.commit()
        return {'msg': f'{len(variantes)} variantes recuperadas'}

    def delete_color(self, db: Session, grupo_id: int, color_id: int, commit: bool = True, representante: Producto = None) -> None:
        if not representante:
            representante = self.repository.get_by_id_including_deleted(db, grupo_id)
        if not representante:
            raise ProductoNoEncontradoException(PRODUCTO_NO_EXISTE)
            
        variantes_inactivas = db.query(Producto).filter(
            Producto.codigo_producto_id == representante.codigo_producto_id,
            Producto.tipo_calzado_id == representante.tipo_calzado_id,
            Producto.material_id == representante.material_id,
            Producto.descripcion == representante.descripcion,
            Producto.color_id == color_id,
            Producto.estado == False
        ).all()

        if not variantes_inactivas:
            raise ProductoNoEncontradoException(PRODUCTO_NO_EXISTE)

        for v in variantes_inactivas:
            self.repository.delete_precios(db, v.id)
            self.repository.delete_imagenes(db, v.id)
            self.repository.delete_producto(db, v.id)
            
        if commit:
            db.commit()

    def update_por_color(
        self,
        db: Session,
        grupo_id: int,
        color_id: int,
        data: ProductoColorUpdate,
    ):
        self._validar_variantes(data.variantes)
        representante = self.repository.get_by_id(db, grupo_id)
        if not representante:
            raise ProductoNoEncontradoException(PRODUCTO_NO_EXISTE)
        
        codigo_producto = self.codigo_repository.get_by_id(db, data.codigo_producto_id)
        if not codigo_producto:
            raise ProductoNoEncontradoException(PRODUCTO_NO_EXISTE)

        variantes_actuales = db.query(Producto).filter(
            Producto.codigo_producto_id == representante.codigo_producto_id,
            Producto.tipo_calzado_id == representante.tipo_calzado_id,
            Producto.material_id == representante.material_id,
            Producto.descripcion == representante.descripcion,
            Producto.color_id == color_id
        ).all()
        
        mapa_existentes = {p.talla_id: p for p in variantes_actuales}
        variantes_entrantes_claves = set()
        
        for variante in data.variantes:
            variantes_entrantes_claves.add(variante.talla_id)

            if variante.talla_id in mapa_existentes:
                producto = mapa_existentes[variante.talla_id]
                # Upsert (Actualizacion in-place)
                producto.tipo_calzado_id = data.tipo_calzado_id
                producto.material_id = data.material_id
                producto.descripcion = data.descripcion
                producto.stock_actual = variante.stock_actual
                producto.stock_minimo = variante.stock_minimo
                producto.stock_maximo = variante.stock_maximo
                producto.estado = variante.estado
                
                # Logica de Precio: 
                precio_actual = None
                for p in producto.precios:
                    if p.estado and p.vigente_hasta is None:
                        precio_actual = p
                        break
                        
                if not precio_actual or precio_actual.precio_compra != variante.precio_compra or precio_actual.precio_venta != variante.precio_venta:
                    if precio_actual:
                        precio_actual.vigente_hasta = datetime.now()
                        precio_actual.estado = False
                        
                    nuevo_precio = PrecioProducto(
                        precio_compra=variante.precio_compra,
                        precio_venta=variante.precio_venta,
                        vigente_desde=datetime.now(),
                        estado=True,
                    )
                    producto.precios.append(nuevo_precio)
            else:
                # Omitir silenciosamente las variantes que no existen
                # Nunca lanzamos error 500 por una combinacion de talla/color inexistente
                pass
                
        # Desactivar variantes que no vinieron en el payload (y que si existian)
        for talla_id, producto in mapa_existentes.items():
            if talla_id not in variantes_entrantes_claves:
                producto.estado = False

        # Procesar actualizaciones de metadata de las imagenes
        if getattr(data, 'imagenes', None):
            for img_data in data.imagenes:
                if img_data.id:
                    imagen = db.query(ProductoImagen).filter(ProductoImagen.id == img_data.id).first()
                    if imagen:
                        imagen.es_principal = img_data.es_principal
                        imagen.orden = img_data.orden

        db.commit()
        return {'msg': 'Producto actualizado correctamente'}

    def bulk_action(self, db: Session, action: str, items: list[dict]) -> dict:
        successful = 0
        failed = 0
        errors = []

        representantes_cache = {}
        for item in items:
            grupo_id = item.get("grupo_id")
            if grupo_id not in representantes_cache:
                try:
                    if action == "desactivar":
                        rep = self.repository.get_by_id(db, grupo_id)
                    else:
                        rep = self.repository.get_by_id_including_deleted(db, grupo_id)
                    
                    if not rep:
                        raise ProductoNoEncontradoException(PRODUCTO_NO_EXISTE)
                    representantes_cache[grupo_id] = rep
                except Exception as e:
                    representantes_cache[grupo_id] = e

        try:
            for item in items:
                grupo_id = item.get("grupo_id")
                color_id = item.get("color_id")
                
                rep_or_err = representantes_cache.get(grupo_id)
                if isinstance(rep_or_err, Exception):
                    raise rep_or_err

                if action == "desactivar":
                    self.desactivar_color(db, grupo_id, color_id, commit=False, representante=rep_or_err)
                elif action == "recuperar":
                    self.recuperar_color(db, grupo_id, color_id, commit=False, representante=rep_or_err)
                elif action == "eliminar":
                    self.delete_color(db, grupo_id, color_id, commit=False, representante=rep_or_err)
                else:
                    raise Exception(f"Acción '{action}' no soportada")
                successful += 1
                
            db.commit()
            
        except Exception as e:
            db.rollback()
            import traceback
            import datetime
            import json
            
            error_details = {
                "timestamp": str(datetime.datetime.now()),
                "exception_type": type(e).__name__,
                "exception_message": str(e),
                "traceback": traceback.format_exc(),
            }
            
            # If it's a SQLAlchemy error, try to extract statement and params
            if hasattr(e, 'statement'):
                error_details['sql_statement'] = getattr(e, 'statement', None)
                error_details['sql_params'] = str(getattr(e, 'params', None))
            if hasattr(e, 'orig'):
                error_details['original_exception'] = str(getattr(e, 'orig', None))
                if hasattr(e.orig, 'pgcode'):
                    error_details['pgcode'] = e.orig.pgcode
                
            try:
                with open("C:/Users/refgu/OneDrive/Documentos/GitHub/InventariosZ/audit_delete.log", "a", encoding="utf-8") as log_file:
                    log_file.write(json.dumps(error_details, indent=2, ensure_ascii=False) + "\n\n")
            except:
                pass
                
            return {
                "successful": 0,
                "failed": len(items),
                "errors": [{"error": f"Operación revertida (Atomicidad). Falló por: {str(e)}"}],
                "message": "Operación masiva fallida y revertida."
            }

        return {
            "successful": successful,
            "failed": 0,
            "errors": [],
            "message": f"Operación masiva completada: {successful} éxitos."
        }
